import openpyxl
from openpyxl import workbook

# GLOBAL VARIABLES
NUMBEROFTEAMS = 12 # constant for number creating dynamic code

PlayerSheet = "SpreadSheets\BowlingLeagueSpreadsheet.xlsx" # Loads the xlsx file from a local relative path to read from
wb_new = openpyxl.load_workbook(PlayerSheet, data_only=True) # loads a new workbook of from the file with the data_only parameter true so formulas in EXCEL are formulated prior to use
sheet_obj = wb_new['Summer2023'] # Gets all the data from the "Summer2023" Sheet in the File


#Start of functions based on the row of individual a player

def GamesPlayed(rowOfPlayer): # Takes in the row of Current Player
    numOfGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3 # Gets the week number from cell V 2
    checker = 0 # Counting the amount of cells that have a 0
    for i in range(numOfGames): # for i in range of amount of games
        currGame = int(sheet_obj.cell(row = rowOfPlayer, column = (23 + i)).value) # setting Current Game Variable to the new row number value
        if currGame == 0: # if the currGame is equal to 0
            checker += 1 # Add one to checker for the current player
    numOfGames -= checker # After the For Loop is over, take the number of games played in league minus the amount of games current player played
    return numOfGames # return amount of games played by this player.

def PlayerAverage(rowOfPlayer): # Takes in the row of Current Player
    numOfGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3 # Gets the week number from cell V 2
    playerTotalPinFall = 0 # Create variable to hold the increment of each number in the Sheet Cells of 'rowOfPlayer' var
    for i in range(numOfGames): # for i in range of amount of games
        CurrGame = int(sheet_obj.cell(row = rowOfPlayer, column = (23 + i)).value) # setting Current Game Variable to the new row number value 
        playerTotalPinFall += CurrGame # Increment PlayerTotalPinfFall By the CurrGame Variable
        if CurrGame == 0: # if currGame is equal to 0
            numOfGames -= 1 # subtract num of games by one, This will account for players whom missed a day of bowling
    if numOfGames == 0: # this is a contingency case for if a player has played 0 games within the first couple weeks so that there is no errors
        pinAverage = 0 # creates pinAverage
        return pinAverage 
    pinAverage = playerTotalPinFall // numOfGames # integer division so it will auto round to nearest Whole Number
    return pinAverage

def PlayerHdcp(rowOfPlayer): #gets the Players Hdcp based on there average
    #gets the player avg from the PlayerAverage function
    avg = PlayerAverage(rowOfPlayer)
    #hdcp is gotten from the (230 - avg) * .8 rounded
    hdcp = ((230 - avg) * 8) // 10
    #return the hdcp
    return hdcp

def TopWeeklyIndividualScore(rowOfPlayer): #Get the Top weekly Individual Score
    #placeholder score that will be replaced
    topScore = 0
    #This is the first game of the week based on the exel sheet
    weeklyGame1 = ((int(sheet_obj.cell(row = 2, column = 22).value) - 1) * 3)
    #range is only 3 because there is 3 games in a week
    for i in range(3):
        #current game is gotten from the exel sheet based on the row of Player and what weekly game it is on
        CurrGame = int(sheet_obj.cell(row = (rowOfPlayer), column = (23 + weeklyGame1)).value)
        #If the top score is less than the Current game than the top score is replaced by the current game
        if topScore < CurrGame:
            topScore = CurrGame
        #this is changing the weekly game by 1 moving it the next game in the week
        weeklyGame1 += 1
    #once it has checked all three games it returns the top game score
    return topScore

def TopIndividualSeriesScratch(rowOfPlayer): #Get Individual Player scratch Series
    #Placeholder that will change later
    seriesScore = 0
    #Sets the current game number to 1 because thats the first game any individual will play
    currentGameNum = 1
    #total number of games is the week number which is gotten by the exel sheet multiplied by 3
    totalNumOfGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3
    #loops everything by the range of the total number of games
    for i in range(totalNumOfGames):
        #placeholder that is reset everytime the loop gets reiterated
        tempSeriesScore = 0
        #the range is 3 because there are only 3 games a week and those three games 
        for i in range(3):
            #current game is gotten from the exel sheet
            currGame = int(sheet_obj.cell(row = (rowOfPlayer), column = (22 + currentGameNum)).value)
            #the current game is added to the temp Series Score
            tempSeriesScore += currGame
            #the current game number is changing by 1 every time the loop gets reiterated
            currentGameNum += 1
        #If the series Score is less than the Temp Series Score then the seriesScore becomes equal to the temp Series Score
        if seriesScore < tempSeriesScore:
            seriesScore = tempSeriesScore
    #once all the looping is finnished then the series Score is returned
    return seriesScore

def TopIndividualScore(rowOfPlayer): #Gets the Top Score seasonal score of a player, top score out of every game played that season
    #number of weeks * 3 because you play three games a week
    numberOfGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3
    #temp score that will be changed
    topScore = 0
    #GameNumber represents what game is being taken at that time
    gameNumber = 1
    #loops through every game and finds the top game played
    for i in range(numberOfGames):
        #CurrGame is equal to the what was scored on each individual game and is changed when the gameNumber gets updated
        CurrGame = int(sheet_obj.cell(row = (rowOfPlayer), column = (22 + gameNumber)).value)
        #if top score is less than CurrGame then topScore becomes the value of CurrGame
        if topScore < CurrGame:
            topScore = CurrGame
        #changing the gameNumber by 1
        gameNumber += 1
    return topScore

def WeeklySeriesScratch(rowOfPlayer): #Get Weekly Series Scratch Score
    #Placeholder Series score that will change
    seriesScore = 0
    #Gets the First game of the week from the exel sheet
    weeklyGame1 = ((int(sheet_obj.cell(row = 2, column = 22).value) - 1) * 3)
    #Range is 3 because there are only three games in a week
    for i in range(3):
        #The current game is gotten from the exel sheet based on the Players row and the weekly Game number
        CurrGame = int(sheet_obj.cell(row = (rowOfPlayer), column = (23 + weeklyGame1)).value)
        #adding the Current game to the series score 
        seriesScore += CurrGame
        #changing the Weekly game by one to get to the next game on the exel sheet
        weeklyGame1 += 1
    #return the Series Score
    return seriesScore

def TotalPinFallAndHdcp(rowOfPlayer): #Gets the total Pinfall and hdcp for a player
    #Getting the maximum amounnt of games that could be played based on what week it is
    totalGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3
    #All of these scores are placeholders that will be used later
    TotalPinfall = 0
    avg = 0
    hdcp = 0
    totalPinFallandHdcp = 0
    currentgame = 0
    #end of placeholders
    #Gets the players last season hdcp from the last season
    bookHdcp = int(sheet_obj.cell(row = rowOfPlayer, column = 19).value)
    #this is a placeholder for how many games the player played
    numberOfGames = 0
    #this is going to go over every game played by the individual
    for i in range(totalGames):
        #increasing the current game by one everytime in the loop
        currentgame += 1
        #getting the current game score from the exel sheet
        currentGameScore = int(sheet_obj.cell(row = rowOfPlayer, column = 22 + currentgame).value)
        #it will go through the if statment if the current game score = 0 because it will mean that they didn't play that game that day
        if currentGameScore != 0:
            #holding the number of games being played by the individual I.E. all non 0 score games
            numberOfGames += 1
            #Everygame that is above 9 and has a remander of 1 when divided by 3 will change the average and that will change the hdcp
            if ((currentgame % 3) == 1) and (numberOfGames > 9):
                avg = TotalPinfall // numberOfGames
                hdcp = ((230 - avg) * 8) // 10
            #If the number of games is under 10 I.E. 9 or less then the hdcp is equal to the bookHdcp  
            elif (numberOfGames < 10):
                hdcp = bookHdcp
            #if number of Games is under 10 then it will use the bookHdcp for the average
            if numberOfGames < 10:
                TotalPinfall += currentGameScore
                totalPinFallandHdcp += currentGameScore + bookHdcp
            #else use the hdcp
            else:
                TotalPinfall += currentGameScore
                totalPinFallandHdcp += currentGameScore + hdcp
    return totalPinFallandHdcp

def TopHdcpGame(rowOfPlayer): #gets the Top Hdcp Game for a player within a season
    #This gets the total weeks from the exel spread sheet
    totalWeeks = int(sheet_obj.cell(row = 2, column = 22).value)
    #Placeholder that will be changed later on and will hold the totalPinFall week by week
    totalPinFall = 0
    #Placeholder that will be returned and shows the Highest hdcp game of the Season so far
    HighHdcpGame = 0
    #the hdcp is set to be the Book Avg to start off with for the first 9 games but it will change later
    hdcp = int(sheet_obj.cell(row = rowOfPlayer, column = 19).value)
    #Placeholder that will change when ever we calculate a new game being played
    numberOfGamesPlayed = 0
    #Loops throught the number of total weeks have been played so far and is controled by currentWeek 
    for currentWeek in range(totalWeeks):
        #Placeholder that is reset every week but holds the high game of the week
        weeklyHighGame = 0
        #this checks if the number of Games Played is greater than or equal to 9, and if so it changes the based on the player preformance of previous weeks
        if (numberOfGamesPlayed >= 9):
            hdcp = ((230 - (totalPinFall // numberOfGamesPlayed)) * 8) // 10
        #This is a loop for each game in any current week based on what the currentWeek is
        for weeklyGameNum in range(3):
            #This gets the game number that is currently being retreaved
            gamenum = 23 + weeklyGameNum + (currentWeek * 3)
            #This gets the score of the game from the exel sheet
            currGame = int(sheet_obj.cell(row = rowOfPlayer, column = gamenum).value)
            #This increases the number of games played which will be used to calculate the average later
            numberOfGamesPlayed += 1
            #This adds the current game to the total pin fall
            totalPinFall += currGame
            #if the current game score is greater than the previous weekly high game then the weekly high game becomes equal to the value
            if currGame > weeklyHighGame:
                weeklyHighGame = currGame
            #else if the value of current game is equal to 0 the number of game played is reduced by 1
            elif currGame == 0:
                numberOfGamesPlayed -=1
        #the hdcp is added to the weeklyHighGame 
        weeklyHighGame += hdcp
        #checks if the weeklyHighGame is greater than the High hdcp game and if so the high hdcp game becomes equal to the weeklyHighGame
        if HighHdcpGame < weeklyHighGame:
            HighHdcpGame = weeklyHighGame
    #once its gone through all the weeks then it returns the HighHdcpGame
    return HighHdcpGame

def TopHdcpSeries(rowOfPlayer): #gets the Top Hdcp Series for a player within a season
    #This gets the total weeks from the exel spread sheet
    totalWeeks = int(sheet_obj.cell(row = 2, column = 22).value)
    #Placeholder that will be changed later on and will hold the totalPinFall week by week
    totalPinFall = 0
    #Placeholder that will be returned and shows the Highest hdcp game of the Season so far
    highHdcpSeries = 0
    #the hdcp is set to be the Book hdcp to start off with for the first 9 games but it will change later
    hdcp = int(sheet_obj.cell(row = rowOfPlayer, column = 19).value) * 3
    #the avg is set to be the book hdcp to start off with for the first 9 games but it will change later
    avg = int(sheet_obj.cell(row = rowOfPlayer, column = 18).value) * 3
    #Placeholder that will change when ever we calculate a new game being played
    numberOfGamesPlayed = 0
    #Loops throught the number of total weeks have been played so far and is controled by currentWeek 
    for currentWeek in range(totalWeeks):
        #Placeholder that is reset every week but holds the high game of the week
        weeklySeriesScore = 0
        #this checks if the number of Games Played is greater than or equal to 9, and if so it changes the based on the player preformance of previous weeks
        if (numberOfGamesPlayed >= 9):
            avg = totalPinFall // numberOfGamesPlayed
            hdcp = ((230 - avg) * 24) // 10
        #This is a loop for each game in any current week based on what the currentWeek is
        for weeklyGameNum in range(3):
            #This gets the game number that is currently being retreaved
            gamenum = 23 + weeklyGameNum + (currentWeek * 3)
            #This gets the score of the game from the exel sheet
            currGame = int(sheet_obj.cell(row = rowOfPlayer, column = gamenum).value)
            #This increases the number of games played which will be used to calculate the average later
            numberOfGamesPlayed += 1
            #This adds the current game to the total pin fall
            totalPinFall += currGame
            weeklySeriesScore += currGame
            #if the current game score is greater than the previous weekly high game then the weekly high game becomes equal to the value
            if currGame > weeklySeriesScore:
                weeklySeriesScore = currGame
            #else if the value of current game is equal to 0 the number of game played is reduced by 1
            elif currGame == 0:
                numberOfGamesPlayed -=1
                weeklySeriesScore += avg
        #the hdcp is added to the weeklyHighGame 
        weeklySeriesScore += hdcp
        #checks if the weeklyHighGame is greater than the High hdcp game and if so the high hdcp game becomes equal to the weeklyHighGame
        if highHdcpSeries < weeklySeriesScore:
            highHdcpSeries = weeklySeriesScore
    #once its gone through all the weeks then it returns the HighHdcpGame
    return highHdcpSeries

def LastWeekHdcp(rowOfPlayer): #Gets the last weeks hdcp
    #Gets the last weeks number from the exel sheet
    lastWeekNum = int(sheet_obj.cell(row = 2, column = 22).value) - 1
    #checks if the week number is 3 or less because if it is then it returns the book avg
    if lastWeekNum <= 3:
        return int(sheet_obj.cell(row = rowOfPlayer, column = 19).value)
    #Placeholder variable that will change later
    totalPinFall = 0
    #total games is last week number * 3 because there are 3 games a week
    totalGames = lastWeekNum * 3
    #placeholder variable that will be changed later
    gamesPlayed = 0
    #A for loop that runs through the loop as many times as there are totalGames
    for gameNum in range(totalGames):
        #Gets the current game score from the exel sheet
        currGame = int(sheet_obj.cell(row = rowOfPlayer, column = 23 + gameNum).value)
        #The current game is added to the totalPinFall
        totalPinFall += currGame
        #one more game is added to gamesPlayed
        gamesPlayed += 1
        #if the current game score is equal to 0 then remove 1 from gamesPlayed
        if currGame == 0:
            gamesPlayed -= 1
    #if the games played is less than 9 then return the book Hdcp
    if gamesPlayed < 9:
        hdcp = int(sheet_obj.cell(row = rowOfPlayer, column = 19).value)
        return hdcp
    #the average is gotten from totalPinFall devided by the number of games played
    avg = totalPinFall // gamesPlayed
    #the hdcp is gotten from the 230 - avg all multiplied by .8
    hdcp = ((230 - avg) * 8) // 10
    return hdcp


#End of functions based on the row of individual a player

#Start of functions based on the team number

def TeamPinFall(teamNumber): # adds up all players pinfall on a team
    teamNumOfGames = int(sheet_obj.cell(row = 2, column = 22).value) * 3 * 3 #Takes the value from cell V 2 which is week number, Multiply it by 3 games a week, and 3 players per team
    firstteamMember = (teamNumber * 3) - 1 # Since the first row of the sheet is made up of organization, we start at 2 and increment by 3 for each team from there. So team number * 3 minus 1
    teamPinFall = 0 # Create variable to hold final value
    for j in range (3): # for j in range 3 (3 being amount of players per team)
        for i in range(teamNumOfGames): # for i in range number of games played by team
            CurrGame = int(sheet_obj.cell(row = (firstteamMember + j), column = (23 + i)).value) # current game equals the value in the new cell
            teamPinFall += CurrGame # adds value from the current Game cell value to teampinfall
    return teamPinFall

def TeamAverage(teamNumber): # team Average method
    firstteamMember = (teamNumber * 3) - 1 # finds the position on the sheet of the firstTeamMember
    teamAverage = 0 # set variable to 0
    for j in range (3): # for j in range of amount of players on a team
        teamAverage += PlayerAverage(firstteamMember + j) # team average is all the players average added together.
    return teamAverage

def TeamTotalPinFallAndHdcp(teamnumber): #Gets the team Total PinFall and Hdcp
    #gets the first team members row on the exel sheet
    teamMemberRow = (3 * teamnumber) - 1
    #placeholder that will change
    teamTotal = 0
    #range is 3 because there are only 3 people in a team
    for i in range(3):
        #player total equals the teamMemberRow + i when it is put through the function TotalPinFallAndHdcp
        playerTotal = TotalPinFallAndHdcp(teamMemberRow + i)
        #player total is added to team total
        teamTotal += playerTotal
    #once all three players 
    return teamTotal

def TopWeeklyTeamGame(teamnumber): #Gets the top weekly game for the team
    #This gets the first game of the week and puts it in the variable currentGameNum
    currentGameNum = (int(sheet_obj.cell(row = 2, column = 22).value) * 3) - 2
    #Placeholder that will be changed later
    highTeamGame = 0
    #range is set to three because there are only 3 games in a week
    for i in range(3):
        #seting and reseting tempTeam game to 0 
        tempTeamGame = 0
        #sets the location on the exel sheet of the row of the first team member
        teamMemberRow = (teamnumber * 3) - 1
        #range is set to three because there are three players on a team
        for i in range(3):
            #it gets the current game based on the team member and the game number 
            currentGame = int(sheet_obj.cell(row = teamMemberRow, column = (22 + currentGameNum)).value) 
            #adds the current game to the tempTeamGame
            tempTeamGame += currentGame
            #changes team member by 1
            teamMemberRow += 1
        #Checks if the Temp game is larger than the High team game and if so sets high team game to equal temp team game
        if tempTeamGame > highTeamGame:
            highTeamGame = tempTeamGame
        #changes the current Game Number by 1
        currentGameNum += 1
    #returns the highTeamGame
    return highTeamGame

def LastWeekTeamHdcp(teamNumber): #gets last weeks team Hdcp
    #gets the first player's row from the exel sheet
    teamPlayerRow = (teamNumber * 3) - 1
    #Gets last weeks number
    lastWeek = int(sheet_obj.cell(row = 2, column = 22).value) - 1
    #gets the number of Games that could be played
    totalHdcp = 0
    for i in range(3):
        numOfGames = lastWeek * 3
        playerTotalPinFall = 0 # Create variable to hold the increment of each number in the Sheet Cells of 'rowOfPlayer' var
        for i in range(lastWeek * 3): # for i in range of amount of games
            CurrGame = int(sheet_obj.cell(row = teamPlayerRow, column = (23 + i)).value) # setting Current Game Variable to the new row number value 
            playerTotalPinFall += CurrGame # Increment PlayerTotalPinfFall By the CurrGame Variable
            if CurrGame == 0: # if currGame is equal to 0
                numOfGames -= 1 # subtract num of games by one, This will account for players whom missed a day of bowling
        if numOfGames <= 9: # this is a contingency case for if a player has played 9 games or less because then you use the book hdcp
            totalHdcp += int(sheet_obj.cell(row = teamPlayerRow, column = (19)).value)
        else:
            totalHdcp += (((230 - (playerTotalPinFall // numOfGames)) * 8) // 10)
    return totalHdcp

def CurrentWeekTeamHdcp(teamNumber):
    #gets the first player's row from the exel sheet
    teamPlayerRow = (teamNumber * 3) - 1
    #Gets last weeks number
    currWeek = int(sheet_obj.cell(row = 2, column = 22).value)
    #gets the number of Games that could be played
    totalHdcp = 0
    for i in range(3):
        numOfGames = currWeek * 3
        playerTotalPinFall = 0 # Create variable to hold the increment of each number in the Sheet Cells of 'rowOfPlayer' var
        for i in range(currWeek * 3): # for i in range of amount of games
            CurrGame = int(sheet_obj.cell(row = teamPlayerRow, column = (23 + i)).value) # setting Current Game Variable to the new row number value 
            playerTotalPinFall += CurrGame # Increment PlayerTotalPinfFall By the CurrGame Variable
            if CurrGame == 0: # if currGame is equal to 0
                numOfGames -= 1 # subtract num of games by one, This will account for players whom missed a day of bowling
        if numOfGames <= 9: # this is a contingency case for if a player has played 9 games or less because then you use the book hdcp
            totalHdcp += int(sheet_obj.cell(row = teamPlayerRow, column = (19)).value)
        else:
            totalHdcp += (((230 - (playerTotalPinFall // numOfGames)) * 8) // 10)
    return totalHdcp

def WeeklyTeamSeries(teamNumber):
    #This is a value holder that will eventualy hold the whole total Series score
    totalSeries = 0
    #this gets the first players row based on the team Number and puts it in terms of exel sheet
    firstPlayerRow = (teamNumber * 3) - 1
    #range is 3 because there are 3 people on a team
    for i in range(3):
        #the weekly Series Scratch score of a single player is added to the total series score
        totalSeries += WeeklySeriesScratch(firstPlayerRow + i)
    #once all three player scratch Series scores are added to total Series, total Series is returned
    return totalSeries

#End of functions based team number

#start of functions Related to top three player/teams for weekly and overall

#sorting is only used in the top three calculations
def sorting(TopThreeScores): #it takes a list and sorts it, high is index 0 and low is index -1 or last in list
    lista = TopThreeScores
    listb = []
    for i in range(3):
        highNum = 0
        for j in lista:
            if j > highNum:
                highNum = j
        listb.append(highNum)
        lista.remove(highNum)
        
    return listb

#Start of Overall games and series

def TopThreeScratchGameMaleAndFemale(numOfTeams): #Gets the top three scratch game scores for Male and Female
    totalNumOfPlayers = numOfTeams * 3
    #Placeholders that will be changed later
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]   
    #the first row where a player shows up in the Exel Sheet 
    rowOfPlayer = 2
    #this runs through the entire list of players in the League
    for i in range(totalNumOfPlayers):
        #Placeholders for Males that will be used to easily swap numbers and names
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #Placeholders for Females that will be used to easily swap numbers and names
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
        #gets the Top Player score for an individual player
        topPlayerScore = TopIndividualScore(rowOfPlayer)
        #gets the gender of the player from the exel sheet, should be m or f
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        #sets the gender to lower to make sure that even if a capital letter was input it would be turned into a lower case
        gender = gender.lower()
        #if the gender is Male then it will be sorted in the male category
        if gender == "m":
            #If the Top Player Score is greater than the lowest number in the top three male scores 
            if topPlayerScore > TopThreeScoresMale[2]:
                #the lowest number is then replaced by the top score
                TopThreeScoresMale[2] = topPlayerScore
                #the player name that had the lowest number in the list is replaced with the current players name
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #score3 which previously held the lowest number and player in the list is being replaced by the new player and score
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #Top Three Scores Male is now being sorted because the new score could be bigger than 2nd place or 1st place
                TopThreeScoresMale = sorting(TopThreeScoresMale)
                #if the new score is actual the new top score then it will sort them accordindly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #else if the new score is actualy the new second place score then it will sort them accordingly
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #else the new score is the third place score and it will be sorted accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]
        #if gender is not male then it is female and will be checked accordingly
        else:
            #if the top player score is greater than the lowest top three score for Females 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #the lowest top three score is replaced by the new topPlayerScore
                TopThreeScoresFemale[2] = topPlayerScore
                #the lowset top three player name is replaced by the new top player score
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #The placeholder that was used for the previous lowest top score is changed to fit the topPlayerScore and the player name
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this is sorting the Top Three Female scores because the new score could be first, second, or third place
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)
                #If the new score is actualy the top score then it will be sorted accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #else if the new score is actualy the 2nd place score then it will be sorted accordingly
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #else the new score is the 3rd place score and it will be sorted accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]             
        #row of Player is being changed by one to get the next player    
        rowOfPlayer += 1
    #Top score is a list that goes male[top score, top player, 2nd score, 2nd player, 3rd score, 3rd player] then it will do the same for female
    TopScoreThenName = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1],
                         TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                         TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenName

def TopThreeScratchSeriesMaleAndFemale(numOfTeams): #Gets the top three scratch Series scores for Male and Female
    totalNumOfPlayers = numOfTeams * 3
    #Placeholders that will be changed later
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]   
    #the first row where a player shows up in the Exel Sheet 
    rowOfPlayer = 2
    #this runs through the entire list of players in the League
    for i in range(totalNumOfPlayers):
        #Placeholders for Males that will be used to easily swap numbers and names
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #Placeholders for Females that will be used to easily swap numbers and names
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
        #gets the Top Player Series for an individual player
        topPlayerScore = TopIndividualSeriesScratch(rowOfPlayer)
        #gets the gender of the player from the exel sheet, should be m or f
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        #sets the gender to lower to make sure that even if a capital letter was input it would be turned into a lower case
        gender = gender.lower()
        #if the gender is Male then it will be sorted in the male category
        if gender == "m":
            #If the Top Player Score is greater than the lowest number in the top three male scores 
            if topPlayerScore > TopThreeScoresMale[2]:
                #the lowest number is then replaced by the top score
                TopThreeScoresMale[2] = topPlayerScore
                #the player name that had the lowest number in the list is replaced with the current players name
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #score3 which previously held the lowest number and player in the list is being replaced by the new player and score
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #Top Three Scores Male is now being sorted because the new score could be bigger than 2nd place or 1st place
                TopThreeScoresMale = sorting(TopThreeScoresMale)
                #if the new score is actual the new top score then it will sort them accordindly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #else if the new score is actualy the new second place score then it will sort them accordingly
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #else the new score is the third place score and it will be sorted accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]
        #if gender is not male then it is female and will be checked accordingly
        else:
            #if the top player score is greater than the lowest top three score for Females 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #the lowest top three score is replaced by the new topPlayerScore
                TopThreeScoresFemale[2] = topPlayerScore
                #the lowset top three player name is replaced by the new top player score
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #The placeholder that was used for the previous lowest top score is changed to fit the topPlayerScore and the player name
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this is sorting the Top Three Female scores because the new score could be first, second, or third place
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)
                #If the new score is actualy the top score then it will be sorted accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #else if the new score is actualy the 2nd place score then it will be sorted accordingly
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #else the new score is the 3rd place score and it will be sorted accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]             
        #row of Player is being changed by one to get the next player    
        rowOfPlayer += 1
    #Top score is a list that goes male[top score, top player, 2nd score, 2nd player, 3rd score, 3rd player] then it will do the same for female
    TopScoreThenName = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1],
                         TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                         TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenName

def TopThreeHdcpGameMaleAndFemale(numOfTeams): #Gets the top three Hdcp games scores for Male and Female
    totalNumOfPlayers = numOfTeams * 3
    #Placeholders that will be changed later
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]   
    #the first row where a player shows up in the Exel Sheet 
    rowOfPlayer = 2
    #this runs through the entire list of players in the League
    for i in range(totalNumOfPlayers):
        #Placeholders for Males that will be used to easily swap numbers and names
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #Placeholders for Females that will be used to easily swap numbers and names
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
        #gets the Top Hdcp Game for an individual player
        topPlayerScore = TopHdcpGame(rowOfPlayer)
        #gets the gender of the player from the exel sheet, should be m or f
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        #sets the gender to lower to make sure that even if a capital letter was input it would be turned into a lower case
        gender = gender.lower()
        #if the gender is Male then it will be sorted in the male category
        if gender == "m":
            #If the Top Player Score is greater than the lowest number in the top three male scores 
            if topPlayerScore > TopThreeScoresMale[2]:
                #the lowest number is then replaced by the top score
                TopThreeScoresMale[2] = topPlayerScore
                #the player name that had the lowest number in the list is replaced with the current players name
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #score3 which previously held the lowest number and player in the list is being replaced by the new player and score
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #Top Three Scores Male is now being sorted because the new score could be bigger than 2nd place or 1st place
                TopThreeScoresMale = sorting(TopThreeScoresMale)
                #if the new score is actual the new top score then it will sort them accordindly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #else if the new score is actualy the new second place score then it will sort them accordingly
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #else the new score is the third place score and it will be sorted accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]
        #if gender is not male then it is female and will be checked accordingly
        else:
            #if the top player score is greater than the lowest top three score for Females 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #the lowest top three score is replaced by the new topPlayerScore
                TopThreeScoresFemale[2] = topPlayerScore
                #the lowset top three player name is replaced by the new top player score
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #The placeholder that was used for the previous lowest top score is changed to fit the topPlayerScore and the player name
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this is sorting the Top Three Female scores because the new score could be first, second, or third place
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)
                #If the new score is actualy the top score then it will be sorted accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #else if the new score is actualy the 2nd place score then it will be sorted accordingly
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #else the new score is the 3rd place score and it will be sorted accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]             
        #row of Player is being changed by one to get the next player    
        rowOfPlayer += 1
    #Top score is a list that goes male[top score, top player, 2nd score, 2nd player, 3rd score, 3rd player] then it will do the same for female
    TopScoreThenName = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1],
                         TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                         TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenName

def TopThreeHdcpSeriesMaleAndFemale(numOfTeams): 
    totalNumOfPlayers = numOfTeams * 3
    #Placeholders that will be changed later
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]   
    #the first row where a player shows up in the Exel Sheet 
    rowOfPlayer = 2
    #this runs through the entire list of players in the League
    for i in range(totalNumOfPlayers):
        #Placeholders for Males that will be used to easily swap numbers and names
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #Placeholders for Females that will be used to easily swap numbers and names
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
        #gets the Top Hdcp Series for an individual player
        topPlayerScore = TopHdcpSeries(rowOfPlayer)
        #gets the gender of the player from the exel sheet, should be m or f
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        #sets the gender to lower to make sure that even if a capital letter was input it would be turned into a lower case
        gender = gender.lower()
        #if the gender is Male then it will be sorted in the male category
        if gender == "m":
            #If the Top Player Score is greater than the lowest number in the top three male scores 
            if topPlayerScore > TopThreeScoresMale[2]:
                #the lowest number is then replaced by the top score
                TopThreeScoresMale[2] = topPlayerScore
                #the player name that had the lowest number in the list is replaced with the current players name
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #score3 which previously held the lowest number and player in the list is being replaced by the new player and score
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #Top Three Scores Male is now being sorted because the new score could be bigger than 2nd place or 1st place
                TopThreeScoresMale = sorting(TopThreeScoresMale)
                #if the new score is actual the new top score then it will sort them accordindly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #else if the new score is actualy the new second place score then it will sort them accordingly
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #else the new score is the third place score and it will be sorted accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]
        #if gender is not male then it is female and will be checked accordingly
        else:
            #if the top player score is greater than the lowest top three score for Females 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #the lowest top three score is replaced by the new topPlayerScore
                TopThreeScoresFemale[2] = topPlayerScore
                #the lowset top three player name is replaced by the new top player score
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #The placeholder that was used for the previous lowest top score is changed to fit the topPlayerScore and the player name
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this is sorting the Top Three Female scores because the new score could be first, second, or third place
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)
                #If the new score is actualy the top score then it will be sorted accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #else if the new score is actualy the 2nd place score then it will be sorted accordingly
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #else the new score is the 3rd place score and it will be sorted accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]             
        #row of Player is being changed by one to get the next player    
        rowOfPlayer += 1
    #Top score is a list that goes male[top score, top player, 2nd score, 2nd player, 3rd score, 3rd player] then it will do the same for female
    TopScoreThenName = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1],
                         TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                         TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenName


#End of Overall games and series

#Start of Weekly games and series

def WeeklyTopThreeScratchMaleAndFemale(numOfTeams): #Gets the Top three weekly Scrach scores for Male and Female
    totalNumOfPlayers = numOfTeams * 3
    #temperary placeholders
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]
    #Row 2 in the exel sheet is the first first person in the exel sheet and will be changed by 
    #+1 later to get other player names and scores and is used in other functions
    rowOfPlayer = 2
    #This loop will make it so that the list goes through every single player and get there top 
    #score and check if its in the top three weekly scores
    for i in range(totalNumOfPlayers):
        #placeholders for each male score then the male player name which will be used to swap names and scores quickly
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #placeholders for female names and scores which will be used to swap names and scores quickly
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]] 

        topPlayerScore = TopWeeklyIndividualScore(rowOfPlayer)
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        gender = gender.lower()
        #This if statement checks for Gender m for male and f for female
        if gender == "m":
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresMale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresMale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresMale = sorting(TopThreeScoresMale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]     
        else:      
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresFemale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]
        #makes it so that the row changes gets the next player in the next loop     
        rowOfPlayer += 1
    #puts in a list where it goes male [top score, top name, 2nd score, 2nd name, 3rd score, 3rd name], then it will do the same for female
    TopScoreThenNameMaleThenFemale = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1], 
                                        TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                                        TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenNameMaleThenFemale

def WeeklyTopThreeHdcpMaleAndFemale(numOfTeams): #Gets the weekly Top three Hdcp games for male and Female
    totalNumOfPlayers = numOfTeams * 3
    #temperary placeholders
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]
    #Row 2 in the exel sheet is the first first person in the exel sheet and will be changed by 
    #+1 later to get other player names and scores and is used in other functions
    rowOfPlayer = 2
    #This loop will make it so that the list goes through every single player and get there top 
    #score and check if its in the top three weekly scores
    for i in range(totalNumOfPlayers):
        #placeholders for each male score then the male player name which will be used to swap names and scores quickly
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #placeholders for female names and scores which will be used to swap names and scores quickly
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]] 

        topPlayerScore = TopWeeklyIndividualScore(rowOfPlayer)
        playerHdcp = LastWeekHdcp(rowOfPlayer)
        topPlayerScore += playerHdcp
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        gender = gender.lower()
        #This if statement checks for Gender m for male and f for female
        if gender == "m":
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresMale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresMale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresMale = sorting(TopThreeScoresMale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]     
        else:      
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresFemale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]
        #makes it so that the row changes gets the next player in the next loop     
        rowOfPlayer += 1
    #puts in a list where it goes male [top score, top name, 2nd score, 2nd name, 3rd score, 3rd name], then it will do the same for female
    TopScoreThenNameMaleThenFemale = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1], 
                                        TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                                        TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenNameMaleThenFemale

def WeeklyTopThreeScratchSeriesMaleAndFemale(numOfTeams): #Get the Top three weekly Scratch series for male and female
    totalNumOfPlayers = numOfTeams * 3
    #temperary placeholders
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]
    #Row 2 in the exel sheet is the first first person in the exel sheet and will be changed by 
    #+1 later to get other player names and scores and is used in other functions
    rowOfPlayer = 2
    #This loop will make it so that the list goes through every single player and get there top 
    #score and check if its in the top three weekly scores
    for i in range(totalNumOfPlayers):
        #placeholders for each male score then the male player name which will be used to swap names and scores quickly
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #placeholders for female names and scores which will be used to swap names and scores quickly
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]] 

        topPlayerScore = WeeklySeriesScratch(rowOfPlayer)
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        gender = gender.lower()
        #This if statement checks for Gender m for male and f for female
        if gender == "m":
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresMale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresMale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresMale = sorting(TopThreeScoresMale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]     
        else:      
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresFemale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]
        #makes it so that the row changes gets the next player in the next loop     
        rowOfPlayer += 1
    #puts in a list where it goes male [top score, top name, 2nd score, 2nd name, 3rd score, 3rd name], then it will do the same for female
    TopScoreThenNameMaleThenFemale = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1], 
                                        TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                                        TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenNameMaleThenFemale

def WeeklyTopThreeHdcpSeriesMaleAndFemale(numOfTeams): #Gets the weekly Top three Hdcp series Male and Female
    totalNumOfPlayers = numOfTeams * 3
    #temperary placeholders
    TopThreeScoresMale = [0,0,0]
    TopThreePlayersMale = ["a","b","c"]
    TopThreeScoresFemale = [0,0,0]
    TopThreePlayersFemale = ["a","b","c"]
    #Row 2 in the exel sheet is the first first person in the exel sheet and will be changed by 
    #+1 later to get other player names and scores and is used in other functions
    rowOfPlayer = 2
    #This loop will make it so that the list goes through every single player and get there top 
    #score and check if its in the top three weekly scores
    for i in range(totalNumOfPlayers):
        #placeholders for each male score then the male player name which will be used to swap names and scores quickly
        score1 = [TopThreeScoresMale[0], TopThreePlayersMale[0]]
        score2 = [TopThreeScoresMale[1], TopThreePlayersMale[1]]
        score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
        #placeholders for female names and scores which will be used to swap names and scores quickly
        score4 = [TopThreeScoresFemale[0], TopThreePlayersFemale[0]]
        score5 = [TopThreeScoresFemale[1], TopThreePlayersFemale[1]]
        score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]] 

        topPlayerScore = WeeklySeriesScratch(rowOfPlayer)
        playerHdcp = LastWeekHdcp(rowOfPlayer) * 3
        topPlayerScore += playerHdcp
        gender = str(sheet_obj.cell(row = rowOfPlayer, column = 5).value)
        gender = gender.lower()
        #This if statement checks for Gender m for male and f for female
        if gender == "m":
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresMale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresMale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersMale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score3 = [TopThreeScoresMale[2], TopThreePlayersMale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresMale = sorting(TopThreeScoresMale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score3[0] == TopThreeScoresMale[0]:
                    TopThreeScoresMale = [score3[0], score1[0], score2[0]]
                    TopThreePlayersMale = [score3[1], score1[1], score2[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score3[0] == TopThreeScoresMale[1]:
                    TopThreeScoresMale = [score1[0], score3[0], score2[0]]
                    TopThreePlayersMale = [score1[1], score3[1], score2[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresMale = [score1[0], score2[0], score3[0]]
                    TopThreePlayersMale = [score1[1], score2[1], score3[1]]     
        else:      
            #if topPlayerScore is greater then the lowest score in the TopThree Scores then it will replace it and sort the 
            if topPlayerScore > TopThreeScoresFemale[2]:
                #changes the lowest score with the new score that is greater than it
                TopThreeScoresFemale[2] = topPlayerScore
                #changes the name of the previously third highest weekly player with the new name that is larger than it
                TopThreePlayersFemale[2] = str(sheet_obj.cell(row = rowOfPlayer, column = 1).value)
                #changes the placeholder with the new values that were just changed
                score6 = [TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
                #this sorts the new list because the new score could be larger than whats in the current largest or second largest position
                TopThreeScoresFemale = sorting(TopThreeScoresFemale)

                #This checks if the new score is the current largest number and changes the lists accordingly
                if score6[0] == TopThreeScoresFemale[0]:
                    TopThreeScoresFemale = [score6[0], score4[0], score5[0]]
                    TopThreePlayersFemale = [score6[1], score4[1], score5[1]]
                #This checks if the new score is currently the second largest number and changes the lists accordingly 
                elif score6[0] == TopThreeScoresFemale[1]:
                    TopThreeScoresFemale = [score4[0], score6[0], score5[0]]
                    TopThreePlayersFemale = [score4[1], score6[1], score5[1]]
                #This means that the new score is the third largest number and changes the lists accordingly
                else:
                    TopThreeScoresFemale = [score4[0], score5[0], score6[0]]
                    TopThreePlayersFemale = [score4[1], score5[1], score6[1]]
        #makes it so that the row changes gets the next player in the next loop     
        rowOfPlayer += 1
    #puts in a list where it goes male [top score, top name, 2nd score, 2nd name, 3rd score, 3rd name], then it will do the same for female
    TopScoreThenNameMaleThenFemale = [TopThreeScoresMale[0], TopThreePlayersMale[0], TopThreeScoresMale[1], TopThreePlayersMale[1], 
                                        TopThreeScoresMale[2], TopThreePlayersMale[2], TopThreeScoresFemale[0], TopThreePlayersFemale[0], 
                                        TopThreeScoresFemale[1], TopThreePlayersFemale[1], TopThreeScoresFemale[2], TopThreePlayersFemale[2]]
    return TopScoreThenNameMaleThenFemale



def WeeklyTeamScratchTop3game(numOfTeams): #Returns a list of 3 teams with teamScore, teamName, teamScore...
    #Placeholder Scores and Names that will be used later
    topTeamScores = [0,0,0]
    topTeamNames = ["a","b","c"]
    #This keeps track of what teamNumber we are referencing on the Exel sheet
    teamNumber = 1
    #It will go through the function for as many teams there are in the league
    for i in range(numOfTeams):
        #Placeholder variables that will hold the top three teams and names that will be used later on to easily transfer data
        team1 = [topTeamScores[0], topTeamNames[0]]
        team2 = [topTeamScores[1], topTeamNames[1]]   
        team3 = [topTeamScores[2], topTeamNames[2]]
        #This gets the top weekly scratch game for a team
        currentTeamTopGame = TopWeeklyTeamGame(teamNumber)
        #gets the current team name from the exel sheet
        teamName = str(sheet_obj.cell(row =(teamNumber * 3) - 1), column = 3)
        #if the current top game is greater than the lowest score in the current top 3 teams then it runs through the if
        if currentTeamTopGame > topTeamScores[2]:
            #the lowest score gets replaced by the new top team game
            topTeamScores[2] = currentTeamTopGame
            #the placeholder team also gets changed accordingly
            team3 = [currentTeamTopGame, teamName]
            #this sorts the new topTeamScores because the new score could be the top score or the second highest
            topTeamScores = sorting(topTeamScores)
            #if the new score is the top score then it orders them accordingly
            if team3[0] == topTeamScores[0]:
                topTeamScores = [team3[0], team1[0], team2[0]]
                topTeamNames = [team3[1], team1[1], team2[1]]
            #if the new score is the Second highest score then it orders it accordingly
            elif team3[0] == topTeamScores[1]:
                topTeamScores = [team1[0], team3[0], team2[0]]
                topTeamNames = [team1[1], team3[1], team2[1]]
            #if the new score is the third highest score then it orders it accordingly
            else:                
                topTeamScores = [team1[0], team2[0], team3[0]]
                topTeamNames = [team1[1], team2[1], team3[1]]
        #changes the team number by +1
        teamNumber += 1
        #puts all the scores and names into a list that goes top score, then top name, then then 2nd score, then 2nd name, 3rd score, 3rd name
    teamScoreThenName = [ topTeamScores[0], topTeamNames[0], 
                          topTeamScores[1], topTeamNames[1], 
                          topTeamScores[2], topTeamNames[2]]
    #returns teamScoreThenName
    return teamScoreThenName

def WeeklyTeamHdcpTop3game(numOfTeams): 
    #Placeholder Scores and Names that will be used later
    topTeamHdcpScores = [0,0,0]
    topTeamNames = ["a","b","c"]
    #This keeps track of what teamNumber we are referencing on the Exel sheet
    teamNumber = 1
    #It will go through the function for as many teams there are in the league
    for i in range(numOfTeams):
        #Placeholder variables that will hold the top three teams and names that will be used later on to easily transfer data
        team1 = [topTeamHdcpScores[0], topTeamNames[0]]
        team2 = [topTeamHdcpScores[1], topTeamNames[1]]   
        team3 = [topTeamHdcpScores[2], topTeamNames[2]]
        #This gets the top weekly scratch game for a team
        currentTeamTopGame = TopWeeklyTeamGame(teamNumber)
        #adds the hdcp from last week
        currentTeamTopGame += LastWeekTeamHdcp(teamNumber)
        #gets the current team name from the exel sheet
        teamName = str(sheet_obj.cell(row =(teamNumber * 3) - 1), column = 3)
        #if the current top game is greater than the lowest score in the current top 3 teams then it runs through the if
        if currentTeamTopGame > topTeamHdcpScores[2]:
            #the lowest score gets replaced by the new top team game
            topTeamHdcpScores[2] = currentTeamTopGame
            #the placeholder team also gets changed accordingly
            team3 = [currentTeamTopGame, teamName]
            #this sorts the new topTeamScores because the new score could be the top score or the second highest
            topTeamHdcpScores = sorting(topTeamHdcpScores)
            #if the new score is the top score then it orders them accordingly
            if team3[0] == topTeamHdcpScores[0]:
                topTeamHdcpScores = [team3[0], team1[0], team2[0]]
                topTeamNames = [team3[1], team1[1], team2[1]]
            #if the new score is the Second highest score then it orders it accordingly
            elif team3[0] == topTeamHdcpScores[1]:
                topTeamHdcpScores = [team1[0], team3[0], team2[0]]
                topTeamNames = [team1[1], team3[1], team2[1]]
            #if the new score is the third highest score then it orders it accordingly
            else:                
                topTeamHdcpScores = [team1[0], team2[0], team3[0]]
                topTeamNames = [team1[1], team2[1], team3[1]]
        #changes the team number by +1
        teamNumber += 1
        #puts all the scores and names into a list that goes top score, then top name, then then 2nd score, then 2nd name, 3rd score, 3rd name
    teamScoreThenName = [ topTeamHdcpScores[0], topTeamNames[0], 
                          topTeamHdcpScores[1], topTeamNames[1], 
                          topTeamHdcpScores[2], topTeamNames[2]]
    #returns teamScoreThenName
    return teamScoreThenName

def WeeklyTeamSeriesTop3Scratch(numOfTeams):
    #Placeholder Scores and Names that will be used later
    topTeamHdcpScores = [0,0,0]
    topTeamNames = ["a","b","c"]
    #This keeps track of what teamNumber we are referencing on the Exel sheet
    teamNumber = 1
    #It will go through the function for as many teams there are in the league
    for i in range(numOfTeams):
        #Placeholder variables that will hold the top three teams and names that will be used later on to easily transfer data
        team1 = [topTeamHdcpScores[0], topTeamNames[0]]
        team2 = [topTeamHdcpScores[1], topTeamNames[1]]   
        team3 = [topTeamHdcpScores[2], topTeamNames[2]]
        #This gets the top weekly scratch series for a team
        currentTeamTopGame = WeeklyTeamSeries(teamNumber)
        #gets the current team name from the exel sheet
        teamName = str(sheet_obj.cell(row =(teamNumber * 3) - 1), column = 3)
        #if the current top score is greater than the lowest score in the current top 3 teams then it runs through the if
        if currentTeamTopGame > topTeamHdcpScores[2]:
            #the lowest score gets replaced by the new top team game
            topTeamHdcpScores[2] = currentTeamTopGame
            #the placeholder team also gets changed accordingly
            team3 = [currentTeamTopGame, teamName]
            #this sorts the new topTeamScores because the new score could be the top score or the second highest
            topTeamHdcpScores = sorting(topTeamHdcpScores)
            #if the new score is the top score then it orders them accordingly
            if team3[0] == topTeamHdcpScores[0]:
                topTeamHdcpScores = [team3[0], team1[0], team2[0]]
                topTeamNames = [team3[1], team1[1], team2[1]]
            #if the new score is the Second highest score then it orders it accordingly
            elif team3[0] == topTeamHdcpScores[1]:
                topTeamHdcpScores = [team1[0], team3[0], team2[0]]
                topTeamNames = [team1[1], team3[1], team2[1]]
            #if the new score is the third highest score then it orders it accordingly
            else:                
                topTeamHdcpScores = [team1[0], team2[0], team3[0]]
                topTeamNames = [team1[1], team2[1], team3[1]]
        #changes the team number by +1
        teamNumber += 1
        #puts all the scores and names into a list that goes top score, then top name, then then 2nd score, then 2nd name, 3rd score, 3rd name
    teamScoreThenName = [ topTeamHdcpScores[0], topTeamNames[0], 
                          topTeamHdcpScores[1], topTeamNames[1], 
                          topTeamHdcpScores[2], topTeamNames[2]]
    #returns teamScoreThenName
    return teamScoreThenName

def WeeklyTeamSeriesTop3Hdcp(numOfTeams):
    #Placeholder Scores and Names that will be used later
    topTeamHdcpScores = [0,0,0]
    topTeamNames = ["a","b","c"]
    #This keeps track of what teamNumber we are referencing on the Exel sheet
    teamNumber = 1
    #It will go through the function for as many teams there are in the league
    for i in range(numOfTeams):
        #Placeholder variables that will hold the top three teams and names that will be used later on to easily transfer data
        team1 = [topTeamHdcpScores[0], topTeamNames[0]]
        team2 = [topTeamHdcpScores[1], topTeamNames[1]]   
        team3 = [topTeamHdcpScores[2], topTeamNames[2]]
        #This gets the top weekly scratch game for a team
        currentTeamTopGame = WeeklyTeamSeries(teamNumber)
        #adds the hdcp from last week
        currentTeamTopGame += LastWeekTeamHdcp(teamNumber) * 3
        #gets the current team name from the exel sheet
        teamName = str(sheet_obj.cell(row =(teamNumber * 3) - 1), column = 3)
        #if the current top game is greater than the lowest score in the current top 3 teams then it runs through the if
        if currentTeamTopGame > topTeamHdcpScores[2]:
            #the lowest score gets replaced by the new top team game
            topTeamHdcpScores[2] = currentTeamTopGame
            #the placeholder team also gets changed accordingly
            team3 = [currentTeamTopGame, teamName]
            #this sorts the new topTeamScores because the new score could be the top score or the second highest
            topTeamHdcpScores = sorting(topTeamHdcpScores)
            #if the new score is the top score then it orders them accordingly
            if team3[0] == topTeamHdcpScores[0]:
                topTeamHdcpScores = [team3[0], team1[0], team2[0]]
                topTeamNames = [team3[1], team1[1], team2[1]]
            #if the new score is the Second highest score then it orders it accordingly
            elif team3[0] == topTeamHdcpScores[1]:
                topTeamHdcpScores = [team1[0], team3[0], team2[0]]
                topTeamNames = [team1[1], team3[1], team2[1]]
            #if the new score is the third highest score then it orders it accordingly
            else:                
                topTeamHdcpScores = [team1[0], team2[0], team3[0]]
                topTeamNames = [team1[1], team2[1], team3[1]]
        #changes the team number by +1
        teamNumber += 1
        #puts all the scores and names into a list that goes top score, then top name, then then 2nd score, then 2nd name, 3rd score, 3rd name
    teamScoreThenName = [ topTeamHdcpScores[0], topTeamNames[0], 
                          topTeamHdcpScores[1], topTeamNames[1], 
                          topTeamHdcpScores[2], topTeamNames[2]]
    #returns teamScoreThenName
    return teamScoreThenName

#End of functions Related to top three player/teams for weekly and overall

# Start of rankings function:

def teamRanks(numOfTeams): #returns a list of Team IDs in order from top scoring to bottom.

    # get all team total points
    # put teamID into a seperate list and code alongside of the points


    teamScores,TeamID = [] # lists for the loops Team Scores will get the team scores and TeamID will hold all the team IDs
    for i in range(1, numOfTeams + 1, 3): # using a 3 in the 3rd parameter of range in order to follow the format of the SpreadSheet
        teamScores.append(int(sheet_obj.cell(row = i, column = 11).value))# get team i score and add to list
        TeamID.append(int(sheet_obj.cell(row = i, column = 2).value)) # get team i ID and add to list

    # small bubble sort method to get team ID in order from top to bottom
    for i in range(numOfTeams-1): 
        for j in range(0,numOfTeams-i-1):
            if teamScores[j] > teamScores[j+1]:
                teamScores[j],teamScores[j+1] = teamScores[j+1], teamScores[j] # swap the team scores
                TeamID[j], TeamID[j+1] = TeamID[j+1], TeamID[j] # swap the correlating team IDs
    #Placeholder that will hold the data of the Team ID then there score in a list of 1 place to last place
    TeamID.reverse()
    return TeamID


#start of Export code:





