'''
This practice file was used in order to get used to reading from one file then outputting to another
'''
import openpyxl
from openpyxl import workbook


PlayerSheet = "SpreadSheets\Player_SpreadSheet.xlsx" # get path to read from this sheet

wb_new = openpyxl.load_workbook(PlayerSheet) # Load this sheet into a variable

sheet_obj = wb_new.active # get all the active cells into this object

tuplePlayers = () # Unused here, plans to use it for top scorers and data management

playerName = '' # Prepare a variable to get player name from cells
cell_obj = '' # prepare a variable to iterate through game scores
row = sheet_obj.max_row # total amount of rows
column = sheet_obj.max_column # total amount of columns

topPlayer = '' # prepare a variable to hold the top player value
topPoint = 0 # used to compare to each top score
currTOT = 0 # will be updated in each iteration and set back to 0 for each player

# starting from row 2, we will iterate and get player name then start going through their games
for j in range(2, row + 1):
    playerName = sheet_obj.cell(row = j, column = 4)
    currTOT = 0
    for i in range(8, column + 1):  # start going through games here
        
        cell_obj = sheet_obj.cell(row = j, column = i) # starting at column i and going until the end of the active sheet we will get the scores
        currTOT += int(cell_obj.value) # turn the string into a variable
        print(playerName.value, cell_obj.value) # print the values of the cells to keep track Not a neccessary line
    
    if currTOT > topPoint: # After each player check to see if they have a higher score
        topPoint = currTOT  # If true get new high score
        topPlayer = playerName # Get new player name
    # print(str(i) + '\n')

print('\n', 'The top player is', topPlayer.value, 'with a total score of', topPoint) #print player name to terminal

PrintSheet = "SpreadSheets\OutputSpreadsheet.xlsx" # Create new path of a spreadsheet to print to

# wb_out = workbook()
# wb_out.save(filename = PrintSheet)

wb_obj = openpyxl.Workbook() # Create new workbook from openpyxl
sheet_out = wb_obj.active # see all active squares

# print top player from last week and their scores
x1 = sheet_out.cell(row = 1, column = 1) # set x1 to the address of A1
x1.value = "Top Player" #set the value at that address

x2 = sheet_out.cell(row = 1, column = 2) # set x1 to the address of A2
x2.value = "Score" #set the value at that address

x3 = sheet_out.cell(row = 2, column = 1) # set x1 to the address of B1
x3.value = topPlayer.value #set the value at that address

x4 = sheet_out.cell(row = 2, column = 2) # set x1 to the address of A2
x4.value = topPoint #set the value at that address


wb_obj.save(PrintSheet) # save the sheet with new values




# Get Player sheet to read from
# Create sheet to print out to

# Compare player values
# Rank players from top to bottom and print to new sheet

# save sheet
