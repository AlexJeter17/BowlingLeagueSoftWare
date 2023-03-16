''' 
This file is for us to practice using openpyxl as our main library. It offers very fundamental functions for grabbing and outputting 
cells from one sheet to data to another sheet.
'''


import openpyxl

# path of the file we are opening from (Relative Path)
path = "SpreadSheets\Player_SpreadSheet.xlsx"

# This loads the Excel sheet into a workbook able to be accessed for python
workbook_obj = openpyxl.load_workbook(path)

# This gets all the active cells form the workbook we have (All cells that are being used)
sheet_obj = workbook_obj.active


# cell_obj = sheet_obj.cell(row = 1, column = 1) # This gets the information from cell A1
# print(cell_obj.value) #print cell
# cell_obj = sheet_obj.cell(row = 1, column = 2) # This gets the information from cell A2
# print(cell_obj.value) #print cell
# cell_obj = sheet_obj.cell(row = 1, column = 3) # This gets the information from cell A3
# print(cell_obj.value) #print cell
# cell_obj = sheet_obj.cell(row = 1, column = 4) # This gets the information from cell A4
# print(cell_obj.value) #print cell


''' Now lets practice reading from multiple cells '''
cell_obj = '' # set up a cell object
row = sheet_obj.max_row # this gains the total amount of rows being used in the sheet
column = sheet_obj.max_column # this gained the toal amount of columns being used in the sheet

# row and column variables hold the amount of rows and columns there are active
print("total rows", row)
print("total columns", column)

# printing from a whole column
print("\nValues of the first column")

for i in range(2, row+1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)

    
    # this loop will go through the second row of the sheet and gain the player scores and output them then finished with 
    # printing out their name and what their average for the previous night was 
    #  
tot = 0
for i in range(3, column-2):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    tot += cell_obj.value
    print(tot)
print("Average for ", sheet_obj.cell(row = 2, column = 2).value, "is", tot / 3)
